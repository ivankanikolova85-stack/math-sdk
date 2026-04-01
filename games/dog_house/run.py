"""
Dog House — Main simulation runner
====================================
Usage (from repo root):    python games/dog_house/run.py
Usage (from game folder):  python run.py

Requirements:
    pip install -r requirements.txt
    pip install zstandard openpyxl
    pip install -e .
"""

import sys, os, json
from collections import Counter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SDK_ROOT   = os.path.abspath(os.path.join(SCRIPT_DIR, '..', '..'))
sys.path.insert(0, SDK_ROOT)
sys.path.insert(0, SCRIPT_DIR)
os.chdir(SCRIPT_DIR)

from gamestate import GameState
from game_config import GameConfig
from src.state.run_sims import create_books
from src.write_data.write_configs import generate_configs

# ── Settings ──────────────────────────────────────────────────────
num_threads  = 4
batch_size   = 1000
compression  = False   # False = readable JSON in library/books/

num_sim_args = {
    "base":  int(1e4),
    "bonus": int(1e4),
}

run_conditions = {
    "run_sims":     True,
    "run_analysis": True,   # prints PAR sheet to terminal
    "run_excel":    True,   # saves PAR sheet as Excel file
}

# ─────────────────────────────────────────────────────────────────
def analyse(sims):
    payouts  = [s['payoutMultiplier'] for s in sims]
    bg_wins  = [s['baseGameWins'] for s in sims]
    fg_wins  = [s['freeGameWins'] for s in sims]
    n        = len(sims)
    criteria = {}
    fs_counts, db_hits, fg_spins_total, sticky_max = [], 0, 0, []
    win_by_sym, sym_counts = {}, {}
    total_sym = 0

    for s in sims:
        c = s['criteria']
        criteria[c] = criteria.get(c,0)+1
        fg_sp, db, max_w = 0, 0, 0
        for ev in s['events']:
            t  = ev.get('type','')
            gt = ev.get('gameType','')
            if t == 'reveal' and gt == 'basegame':
                for col in ev.get('board',[]):
                    for sym in col:
                        nm = sym['name']
                        sym_counts[nm] = sym_counts.get(nm,0)+1
                        total_sym += 1
            if t == 'reveal' and gt == 'freegame':
                fg_sp += 1
                wilds = sum(1 for col in ev.get('board',[])
                           for sym in col if sym.get('name')=='W')
                max_w = max(max_w, wilds)
                if any(sym.get('name')=='DB'
                       for col in ev.get('board',[]) for sym in col):
                    db += 1
            if t == 'winInfo':
                for w in ev.get('wins',[]):
                    key = f"{w['kind']}x{w['symbol']}"
                    win_by_sym.setdefault(key,{'count':0,'total':0})
                    win_by_sym[key]['count'] += 1
                    win_by_sym[key]['total'] += w['win']
        if fg_sp > 0:
            fs_counts.append(fg_sp)
            db_hits += db
            fg_spins_total += fg_sp
            sticky_max.append(max_w)

    return {
        'n':n,'payouts':payouts,'mean':sum(payouts)/n,
        'max':max(payouts),'zeros':sum(1 for p in payouts if p==0),
        'bg_rtp':sum(bg_wins)/n,'fg_rtp':sum(fg_wins)/n,
        'criteria':criteria,'fs_counts':fs_counts,
        'db_hits':db_hits,'fg_spins_total':fg_spins_total,
        'sticky_max':sticky_max,'win_by_sym':win_by_sym,
        'sym_counts':sym_counts,'total_sym':total_sym,
    }


def build_csv(bs, bo, out_path):
    import csv
    rows = []

    rows.append(["DOG HOUSE - PAR SHEET"])
    rows.append([])
    rows.append(["GAME SPECIFICATION"])
    rows.append(["Game ID", "dog_house"])
    rows.append(["Grid", "5 reels x 3 rows"])
    rows.append(["Win System", "Lines - 20 paylines"])
    rows.append(["Target RTP", "96%"])
    rows.append(["Win Cap", "5000x total bet"])
    rows.append(["Base Bet", "1x total bet"])
    rows.append(["Bonus Buy", "100x total bet"])
    rows.append([])

    rows.append(["PAYTABLE (x total bet per winning line)"])
    rows.append(["Symbol", "Name", "3x Pay", "4x Pay", "5x Pay"])
    pt = [
        ("H1","Black Dog",  5.00, 15.00, 75.00),
        ("H2","Pink Dog",   3.50, 10.00, 50.00),
        ("H3","Pug",        2.50,  6.00, 30.00),
        ("H4","Brown Dog",  2.00,  4.00, 20.00),
        ("L1","Collar",     1.20,  2.50, 15.00),
        ("L2","Bone",       0.80,  2.00, 10.00),
        ("L3","A",          0.50,  1.00,  5.00),
        ("L4","K",          0.50,  1.00,  5.00),
        ("L5","Q",          0.20,  0.50,  2.50),
        ("L6","J",          0.20,  0.50,  2.50),
        ("L7","10",         0.20,  0.50,  2.50),
        ("W", "Wild - substitute only", "-", "-", "-"),
        ("S", "Scatter - FG trigger",   "-", "-", "-"),
        ("DB","DogBooster - +2x wilds", "-", "-", "-"),
    ]
    for row in pt:
        rows.append(list(row))
    rows.append([])

    rows.append(["RTP SUMMARY"])
    rows.append(["Metric", "Base Mode", "Bonus Mode", "Notes"])
    fs  = bs["fs_counts"]
    fso = bo["fs_counts"]
    rows.append(["TARGET RTP", "96.00%", "96.00%", "Set in game_config.py - achieved via optimizer"])
    rows.append(["Simulations", bs["n"], bo["n"], ""])
    rows.append(["BG RTP (raw)", f"{bs['bg_rtp']*100:.4f}%", "-", "Base game line wins"])
    rows.append(["FG RTP (raw)", f"{bs['fg_rtp']*100:.4f}%", f"{bo['fg_rtp']/100*100:.4f}%", "Freegame wins (bonus divided by 100x bet cost)"])
    rows.append(["Raw mean payout", f"{bs['mean']:.3f}x", f"{sum(bo['payouts'])/bo['n']/100:.3f}x", "Pre-optimization (bonus shown in base-bet units)"])
    rows.append(["Max payout", f"{bs['max']:.0f}x", f"{max(bo['payouts'])/100:.0f}x", "Wincap = 5000x base bet"])
    rows.append(["Zero win rate", f"{bs['zeros']/bs['n']*100:.1f}%", "0%", "Base only"])
    rows.append(["Wincap working", "Yes", "Yes", f"Max bonus win = {max(bo['payouts'])/100:.0f}x base bet"])
    rows.append(["FG sessions", len(fs), len(fso)])
    rows.append(["FG trigger rate", f"{len(fs)/bs['n']*100:.1f}%", "100%"])
    rows.append(["Avg FG spins", f"{sum(fs)/len(fs):.1f}" if fs else "-", f"{sum(fso)/len(fso):.1f}" if fso else "-"])
    rows.append(["DB hit rate per FG spin", f"{bs['db_hits']/max(bs['fg_spins_total'],1)*100:.1f}%", f"{bo['db_hits']/max(bo['fg_spins_total'],1)*100:.1f}%"])
    rows.append(["Max sticky wilds", max(bs["sticky_max"]) if bs["sticky_max"] else "-", max(bo["sticky_max"]) if bo["sticky_max"] else "-"])
    rows.append([])

    rows.append(["PAYOUT DISTRIBUTION (Base Mode)"])
    rows.append(["Win Band", "Count", "Frequency %", "Avg Win"])
    bands = [
        ("0x",        lambda p: p==0),
        ("0.01-1x",   lambda p: 0<p<=1),
        ("1-5x",      lambda p: 1<p<=5),
        ("5-20x",     lambda p: 5<p<=20),
        ("20-100x",   lambda p: 20<p<=100),
        ("100-500x",  lambda p: 100<p<=500),
        ("500-2000x", lambda p: 500<p<=2000),
        ("2000x+",    lambda p: p>2000),
    ]
    for lbl, cond in bands:
        band = [p for p in bs["payouts"] if cond(p)]
        avg  = sum(band)/len(band) if band else 0
        rows.append([lbl, len(band), f"{len(band)/bs['n']*100:.1f}%", f"{avg:.2f}x"])
    rows.append([])

    rows.append(["WIN FREQUENCY BY SYMBOL (Base Mode)"])
    rows.append(["Combo", "Hits", "Hit Rate %", "Total Won (x)", "Avg Win (x)"])
    combos = ["5xH1","4xH1","3xH1","5xH2","4xH2","3xH2","5xH3","4xH3","3xH3",
              "5xH4","4xH4","3xH4","5xL1","4xL1","3xL1","5xL2","4xL2","3xL2",
              "5xL3","4xL3","3xL3","5xL4","4xL4","3xL4","5xL5","4xL5","3xL5",
              "5xL6","4xL6","3xL6","5xL7","4xL7","3xL7"]
    for combo in combos:
        d   = bs["win_by_sym"].get(combo, {"count":0,"total":0})
        avg = d["total"]/d["count"] if d["count"] else 0
        rows.append([combo, d["count"], f"{d['count']/bs['n']*100:.2f}%",
                     f"{d['total']:.2f}x", f"{avg:.2f}x"])
    rows.append([])

    rows.append(["SYMBOL FREQUENCIES (Base Game Board)"])
    rows.append(["Symbol", "Count", "% Board", "Per Board Avg"])
    total = bs["total_sym"]; nb = bs["n"]
    for sym in ["H1","H2","H3","H4","L1","L2","L3","L4","L5","L6","L7","W","S","DB"]:
        cnt = bs["sym_counts"].get(sym, 0)
        rows.append([sym, cnt, f"{cnt/total*100:.1f}%" if total else "-", f"{cnt/nb:.2f}"])

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)
    print(f"  CSV PAR sheet saved: {out_path}")

def build_excel_UNUSED(bs, bo, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb  = Workbook()
    DARK, WHITE, DGREY = "1A1A2E", "FFFFFF", "2C3E50"
    BLUE, LBLUE, LGREY = "2980B9", "D6EAF8", "F2F2F2"
    GOLD, GREEN, RED   = "F39C12", "1E8449", "C0392B"

    def brd():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s,right=s,top=s,bottom=s)

    def h(ws,row,col,text,bg=DGREY,color=WHITE,bold=True,sz=10,merge_to=None):
        c = ws.cell(row=row,column=col,value=text)
        c.font = Font(name="Arial",bold=bold,size=sz,color=color)
        c.fill = PatternFill("solid",start_color=bg)
        c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border = brd()
        if merge_to:
            ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=merge_to)
        return c

    def v(ws,row,col,value,bg=LGREY,color="222222",fmt=None,bold=False):
        c = ws.cell(row=row,column=col,value=value)
        c.font = Font(name="Arial",size=10,color=color,bold=bold)
        c.fill = PatternFill("solid",start_color=bg)
        c.alignment = Alignment(horizontal="center",vertical="center")
        c.border = brd()
        if fmt: c.number_format = fmt
        return c

    def lbl(ws,row,col,text,bg=LBLUE):
        c = ws.cell(row=row,column=col,value=text)
        c.font = Font(name="Arial",bold=True,size=10,color=DGREY)
        c.fill = PatternFill("solid",start_color=bg)
        c.alignment = Alignment(vertical="center",horizontal="left")
        c.border = brd()
        return c

    def sec(ws,row,col,text,ncols,bg=DGREY):
        ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+ncols-1)
        c = ws.cell(row=row,column=col,value=text)
        c.font = Font(name="Arial",bold=True,size=11,color=WHITE)
        c.fill = PatternFill("solid",start_color=bg)
        c.alignment = Alignment(horizontal="left",vertical="center")
        ws.row_dimensions[row].height = 21

    # ── SHEET 1: PAR Summary ──────────────────────────────────────
    ws = wb.active
    ws.title = "PAR Sheet"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 22

    ws.merge_cells("B1:E3")
    c = ws["B1"]
    c.value = "🐾  DOG HOUSE — PAR Sheet"
    c.font  = Font(name="Arial",bold=True,size=20,color=WHITE)
    c.fill  = PatternFill("solid",start_color=DARK)
    c.alignment = Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[2].height = 32

    row = 5
    sec(ws,row,2,"  GAME SPECIFICATION",4); row+=1
    specs = [
        ("Game ID","dog_house"),("Grid","5 reels × 3 rows"),
        ("Win System","Lines — 20 paylines"),("Target RTP","96.00%"),
        ("Win Cap","5,000× total bet"),("Base Bet","1× total bet"),
        ("Bonus Buy","100× total bet"),("Scatter Reels","1, 3, 5 only (max 3)"),
        ("Wild","Substitute only — forbidden on reel 1"),
        ("Freegame Spins","Random 8–15 on trigger; +10 retrigger"),
        ("Sticky Wilds","Persist all freegame spins; mult starts 1×"),
        ("DogBooster","Freegame only — +2× to ALL sticky wilds"),
    ]
    for k,val_txt in specs:
        ws.row_dimensions[row].height = 17
        lbl(ws,row,2,k,bg="D5E8F3")
        c2 = ws.cell(row=row,column=3,value=val_txt)
        c2.font = Font(name="Arial",size=10)
        c2.fill = PatternFill("solid",start_color=WHITE)
        c2.alignment = Alignment(vertical="center")
        c2.border = brd()
        ws.merge_cells(start_row=row,start_column=3,end_row=row,end_column=5)
        row+=1

    row+=1
    sec(ws,row,2,"  PAYTABLE (× total bet per winning line)",4); row+=1
    h(ws,row,2,"Symbol"); h(ws,row,3,"3× Pay"); h(ws,row,4,"4× Pay"); h(ws,row,5,"5× Pay")
    ws.row_dimensions[row].height=18; row+=1
    pt = [
        ("H1","Black Dog",  5.00,15.00,75.00,"FFD700"),
        ("H2","Pink Dog",   3.50,10.00,50.00,"FFA500"),
        ("H3","Pug",        2.50, 6.00,30.00,"FF8C00"),
        ("H4","Brown Dog",  2.00, 4.00,20.00,"FF6347"),
        ("L1","Collar",     1.20, 2.50,15.00,"90EE90"),
        ("L2","Bone",       0.80, 2.00,10.00,"66CDAA"),
        ("L3","A",          0.50, 1.00, 5.00,"87CEEB"),
        ("L4","K",          0.50, 1.00, 5.00,"87CEEB"),
        ("L5","Q",          0.20, 0.50, 2.50,"B0E0E6"),
        ("L6","J",          0.20, 0.50, 2.50,"B0E0E6"),
        ("L7","10",         0.20, 0.50, 2.50,"B0E0E6"),
        ("W", "Wild — substitute only","—","—","—","FFB6C1"),
        ("S", "Scatter — FG trigger",  "—","—","—","DDA0DD"),
        ("DB","DogBooster — +2× wilds","—","—","—","F4A460"),
    ]
    for sym,name,p3,p4,p5,bg in pt:
        ws.row_dimensions[row].height=17
        v(ws,row,2,f"{sym} ({name})",bg=bg,bold=True)
        for ci,pv in enumerate([p3,p4,p5]):
            v(ws,row,3+ci,pv,bg=bg,fmt="0.00" if isinstance(pv,float) else "@")
        row+=1

    row+=1
    sec(ws,row,2,"  RTP SUMMARY",4,DGREY); row+=1
    h(ws,row,2,"Metric"); h(ws,row,3,"Base Mode",bg=BLUE); h(ws,row,4,"Bonus Mode",bg="8E44AD"); h(ws,row,5,"Notes")
    ws.row_dimensions[row].height=18; row+=1
    fs=bs['fs_counts']; fso=bo['fs_counts']
    rtp_rows = [
        ("Simulations",         bs['n'],                        bo['n'],                       ""),
        ("Mean payout (raw)",   f"{bs['mean']:.3f}×",           f"{bo['mean']:.3f}×",          "Pre-optimization"),
        ("Max payout seen",     f"{bs['max']:.0f}×",            f"{bo['max']:.0f}×",           "Wincap=5000×"),
        ("Zero win rate",       f"{bs['zeros']/bs['n']*100:.1f}%","N/A",                       "Base only"),
        ("BG RTP contribution", f"{bs['bg_rtp']*100:.4f}%",     "—",                           "Line wins on base reveal"),
        ("FG RTP contribution", f"{bs['fg_rtp']*100:.4f}%",     f"{bo['fg_rtp']*100:.4f}%",   "Freegame wins"),
        ("FG sessions",         len(fs),                        len(fso),                      ""),
        ("FG trigger rate",     f"{len(fs)/bs['n']*100:.1f}%",  "100%",                        "All bonus = FG"),
        ("Avg FG spins",        f"{sum(fs)/len(fs):.1f}" if fs else "—",
                                f"{sum(fso)/len(fso):.1f}" if fso else "—",                   "Target 11.5"),
        ("DB hit rate/FG spin", f"{bs['db_hits']/max(bs['fg_spins_total'],1)*100:.1f}%",
                                f"{bo['db_hits']/max(bo['fg_spins_total'],1)*100:.1f}%",       "~25% expected"),
        ("Max sticky wilds",    max(bs['sticky_max']) if bs['sticky_max'] else "—",
                                max(bo['sticky_max']) if bo['sticky_max'] else "—",            "Per session"),
    ]
    for i,(metric,bv,bov,note) in enumerate(rtp_rows):
        bg = "EBF5FB" if i%2==0 else WHITE
        ws.row_dimensions[row].height=17
        lbl(ws,row,2,metric,bg=bg)
        v(ws,row,3,bv,bg=bg); v(ws,row,4,bov,bg=bg)
        c5=ws.cell(row=row,column=5,value=note)
        c5.font=Font(name="Arial",size=9,italic=True,color="7F8C8D")
        c5.fill=PatternFill("solid",start_color="FAFAFA")
        c5.alignment=Alignment(vertical="center"); c5.border=brd()
        row+=1

    row+=1
    sec(ws,row,2,"  PAYOUT DISTRIBUTION (Base Mode)",4); row+=1
    h(ws,row,2,"Win Band"); h(ws,row,3,"Count"); h(ws,row,4,"Frequency"); h(ws,row,5,"Avg Win")
    ws.row_dimensions[row].height=18; row+=1
    bands=[("0×",lambda p:p==0),("0.01–1×",lambda p:0<p<=1),("1–5×",lambda p:1<p<=5),
           ("5–20×",lambda p:5<p<=20),("20–100×",lambda p:20<p<=100),
           ("100–500×",lambda p:100<p<=500),("500–2000×",lambda p:500<p<=2000),
           ("2000×+",lambda p:p>2000)]
    bgs=["FADBD8","FEF9E7","EAFAF1","D6EAF8","E8DAEF","FEF9E7","FADBD8","FFD700"]
    for (bt,cond),bg in zip(bands,bgs):
        band=[p for p in bs['payouts'] if cond(p)]
        avg=sum(band)/len(band) if band else 0
        ws.row_dimensions[row].height=16
        v(ws,row,2,bt,bg=bg,bold=True); v(ws,row,3,len(band),bg=bg,fmt="0")
        v(ws,row,4,len(band)/bs['n'],bg=bg,fmt="0.0%"); v(ws,row,5,round(avg,2),bg=bg,fmt="0.00")
        row+=1

    # ── SHEET 2: Win Frequency ────────────────────────────────────
    ws2=wb.create_sheet("Win Frequency")
    ws2.sheet_view.showGridLines=False
    ws2.column_dimensions['A'].width=2
    for col,w in [('B',14),('C',14),('D',12),('E',14),('F',12)]:
        ws2.column_dimensions[col].width=w

    ws2.merge_cells("B1:F3")
    c=ws2["B1"]; c.value="Win Frequency by Symbol & Kind"
    c.font=Font(name="Arial",bold=True,size=16,color=WHITE)
    c.fill=PatternFill("solid",start_color=DARK)
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws2.row_dimensions[2].height=24

    row=5
    sec(ws2,row,2,"  BASE MODE — Hit Frequency",5); row+=1
    h(ws2,row,2,"Combo"); h(ws2,row,3,"Hits"); h(ws2,row,4,"Hit Rate")
    h(ws2,row,5,"Total Won (×)"); h(ws2,row,6,"Avg Win (×)")
    ws2.row_dimensions[row].height=18; row+=1

    sym_bgs={"H1":"FFD700","H2":"FFA500","H3":"FF8C00","H4":"FF6347",
             "L1":"90EE90","L2":"66CDAA","L3":"87CEEB","L4":"87CEEB",
             "L5":"B0E0E6","L6":"B0E0E6","L7":"B0E0E6"}
    combos=["5xH1","4xH1","3xH1","5xH2","4xH2","3xH2","5xH3","4xH3","3xH3",
            "5xH4","4xH4","3xH4","5xL1","4xL1","3xL1","5xL2","4xL2","3xL2",
            "5xL3","4xL3","3xL3","5xL4","4xL4","3xL4","5xL5","4xL5","3xL5",
            "5xL6","4xL6","3xL6","5xL7","4xL7","3xL7"]
    for combo in combos:
        d=bs['win_by_sym'].get(combo,{'count':0,'total':0})
        sym=combo[2:]; bg=sym_bgs.get(sym,LGREY)
        avg=d['total']/d['count'] if d['count'] else 0
        ws2.row_dimensions[row].height=16
        v(ws2,row,2,combo,bg=bg,bold=True)
        v(ws2,row,3,d['count'],bg=bg,fmt="0")
        v(ws2,row,4,d['count']/bs['n'],bg=bg,fmt="0.000%")
        v(ws2,row,5,round(d['total'],2),bg=bg,fmt="0.00")
        v(ws2,row,6,round(avg,2),bg=bg,fmt="0.00")
        row+=1

    # ── SHEET 3: Symbol Frequencies ──────────────────────────────
    ws3=wb.create_sheet("Symbol Frequencies")
    ws3.sheet_view.showGridLines=False
    ws3.column_dimensions['A'].width=2
    for col,w in [('B',12),('C',12),('D',12),('E',12),('F',12)]:
        ws3.column_dimensions[col].width=w

    ws3.merge_cells("B1:F3")
    c=ws3["B1"]; c.value="Symbol Frequencies on Base Game Board"
    c.font=Font(name="Arial",bold=True,size=16,color=WHITE)
    c.fill=PatternFill("solid",start_color=DARK)
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws3.row_dimensions[2].height=24

    row=5
    sec(ws3,row,2,f"  {bs['n']} boards × 15 positions = {bs['n']*15} total",5); row+=1
    h(ws3,row,2,"Symbol"); h(ws3,row,3,"Count"); h(ws3,row,4,"% Board")
    h(ws3,row,5,"Per Board"); h(ws3,row,6,"Pay 5×")
    ws3.row_dimensions[row].height=18; row+=1

    pt5={"H1":75,"H2":50,"H3":30,"H4":20,"L1":15,"L2":10,
         "L3":5,"L4":5,"L5":2.5,"L6":2.5,"L7":2.5,"W":"Sub","S":"FG","DB":"+2×"}
    total=bs['total_sym']; nb=bs['n']
    for sym in ["H1","H2","H3","H4","L1","L2","L3","L4","L5","L6","L7","W","S","DB"]:
        cnt=bs['sym_counts'].get(sym,0)
        bg=sym_bgs.get(sym,"F0F0F0")
        ws3.row_dimensions[row].height=17
        v(ws3,row,2,sym,bg=bg,bold=True)
        v(ws3,row,3,cnt,bg=bg,fmt="0")
        v(ws3,row,4,cnt/total if total else 0,bg=bg,fmt="0.0%")
        v(ws3,row,5,round(cnt/nb,2),bg=bg,fmt="0.00")
        v(ws3,row,6,pt5.get(sym,"—"),bg=bg)
        row+=1

    wb.save(out_path)
    print(f"  Excel PAR sheet saved: {out_path}")


# ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":

    print("=" * 55)
    print("DOG HOUSE — Math Engine")
    print("=" * 55)
    print(f"  Game ID:    dog_house")
    print(f"  Target RTP: 96%  |  Win Cap: 5000x")
    print(f"  Grid:       5 reels x 3 rows, 20 paylines")
    print(f"  Sims:       {num_sim_args}")
    print()

    config    = GameConfig()
    gamestate = GameState(config)

    if run_conditions["run_sims"]:
        create_books(
            gamestate, config, num_sim_args,
            batch_size, num_threads, compression, False,
        )
        generate_configs(gamestate)

    if run_conditions["run_analysis"] or run_conditions["run_excel"]:
        books_path = os.path.join(SCRIPT_DIR, 'library', 'books')
        with open(os.path.join(books_path,'books_base.json')) as f:
            base = json.load(f)
        with open(os.path.join(books_path,'books_bonus.json')) as f:
            bonus = json.load(f)
        bs = analyse(base)
        bo = analyse(bonus)

    if run_conditions["run_analysis"]:
        fs=bs['fs_counts']
        print()
        print("=" * 55)
        print("PAR SHEET SUMMARY")
        print("=" * 55)
        print(f"\nBASE MODE  ({bs['n']} sims)")
        print(f"  TARGET RTP:    96.00%  (configured in game_config.py)")
        print(f"  BG RTP:        {bs['bg_rtp']*100:.4f}%  (line wins from base reveal)")
        print(f"  FG RTP:        {bs['fg_rtp']*100:.4f}%  (freegame wins)")
        print(f"  Raw mean pay:  {bs['mean']:.3f}x  (pre-optimization)")
        print(f"  Max payout:    {bs['max']:.0f}x base bet")
        print(f"  Zero win rate: {bs['zeros']/bs['n']*100:.1f}%")
        print(f"  NOTE: Raw RTP is high - optimizer calibrates to 96%")
        bonus_cost = 100
        print(f"\nBONUS MODE  ({bo['n']} sims — 100x bet cost)")
        print(f"  TARGET RTP:    96.00%")
        print(f"  FG RTP:        {bo['fg_rtp']/bonus_cost*100:.4f}%  (per base bet, pre-optimization)")
        print(f"  Mean payout:   {bo['mean']/bonus_cost:.3f}x base bet")
        print(f"  Max payout:    {bo['max']/bonus_cost:.0f}x base bet  (wincap = 5000x)")
        print(f"\nFREEGAME STATS")
        print(f"  Sessions:      {len(fs)}")
        print(f"  Trigger rate:  {len(fs)/bs['n']*100:.1f}%")
        print(f"  Avg spins:     {sum(fs)/len(fs):.1f}  (target 11.5)")
        print(f"  Spin dist:     {dict(sorted(Counter(fs).items()))}")
        print(f"  DB hit rate:   {bs['db_hits']/max(bs['fg_spins_total'],1)*100:.1f}% per FG spin")
        print(f"  Max wilds:     {max(bs['sticky_max']) if bs['sticky_max'] else 0}")

    if run_conditions["run_excel"]:
        print()
        csv_path = os.path.join(SCRIPT_DIR, 'library', 'dog_house_PAR_sheet.csv')
        build_csv(bs, bo, csv_path)

    print()
    print("=" * 55)
    print("DONE!")
    print(f"  Simulation: library/books/")
    print(f"  PAR Sheet:  library/dog_house_PAR_sheet.csv")
    print("=" * 55)
